"""
Экспорт данных компаний в Excel.
Корректная запись всех полей, URL как гиперссылки.
"""
import logging
from typing import List
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Company

logger = logging.getLogger(__name__)


class ExcelExporter:
    def __init__(self):
        self.workbook = None
        self.worksheet = None

    def export_to_excel(self, companies: List[Company], filename: str) -> str:
        if not companies:
            raise ValueError("Список компаний пуст")

        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Компании 2GIS"

        headers = [
            'Название компании',
            'Город',
            'Телефон',
            'Адрес',
            'Рейтинг',
            'Количество голосов',
            'Информация',
            'Ссылка'
        ]
        header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, h in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, company in enumerate(companies, start=2):
            name = (company.name or '').strip() or '—'
            city = (company.city or '').strip() or '—'
            phone = (company.phone or '').strip() or '—'
            address = (company.address or '').strip() or '—'
            rating_str = str(company.rating) if company.rating is not None else '—'
            voters_str = str(company.voters_count) if company.voters_count is not None else '—'
            info = (company.info or '').strip() or '—'
            url = (company.url or '').strip() or ''

            self.worksheet.cell(row=row_idx, column=1, value=str(name)[:500])
            self.worksheet.cell(row=row_idx, column=2, value=str(city)[:80])
            self.worksheet.cell(row=row_idx, column=3, value=str(phone)[:100])
            self.worksheet.cell(row=row_idx, column=4, value=str(address)[:500])
            self.worksheet.cell(row=row_idx, column=5, value=rating_str)
            self.worksheet.cell(row=row_idx, column=6, value=voters_str)
            self.worksheet.cell(row=row_idx, column=7, value=str(info)[:1000])
            link_cell = self.worksheet.cell(row=row_idx, column=8)
            if url and url.startswith('http'):
                # Короткий текст вместо длинного URL — гиперссылка работает при клике
                link_cell.hyperlink = url.split('?')[0]  # Убираем query-параметры для стабильности
                link_cell.value = "Открыть"
                link_cell.font = Font(color="0563C1", underline="single")
            else:
                link_cell.value = url or '—'

            for col in range(1, 9):
                self.worksheet.cell(row=row_idx, column=col).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

        self.worksheet.column_dimensions['A'].width = 35
        self.worksheet.column_dimensions['B'].width = 18
        self.worksheet.column_dimensions['C'].width = 38
        self.worksheet.column_dimensions['D'].width = 45
        self.worksheet.column_dimensions['E'].width = 10
        self.worksheet.column_dimensions['F'].width = 15
        self.worksheet.column_dimensions['G'].width = 50
        self.worksheet.column_dimensions['H'].width = 12

        filepath = self._ensure_filepath(filename)
        self.workbook.save(filepath)
        logger.info(f"Файл сохранен: {filepath}")
        return str(filepath)

    def _ensure_filepath(self, filename: str) -> Path:
        filepath = Path(filename)
        if not filepath.suffix:
            filepath = filepath.with_suffix('.xlsx')
        if not filepath.is_absolute():
            filepath = Path.cwd() / filepath
        filepath.parent.mkdir(parents=True, exist_ok=True)
        return filepath
